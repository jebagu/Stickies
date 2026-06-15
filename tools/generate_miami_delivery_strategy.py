from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path("/Users/jeremyguillory/Downloads/SS Miami delivery strategy.xlsx")
OUTPUT = Path(
    "/Users/jeremyguillory/Documents/vibecode-projects/SS React Flow Charts/"
    "ss-miami-delivery-strategy-flowchart.json"
)


def slug(prefix: str, value: str) -> str:
    normalized = str(value).strip().lower().replace("&", "and")
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return f"{prefix}_{normalized}" if prefix else normalized


def initials(name: str) -> str:
    if re.fullmatch(r"[A-Z0-9]+", name.strip()):
        return name.strip()[:4]

    parts = [
        part
        for part in re.sub(r"[,.]", " ", name).split()
        if part.lower() != "van"
    ]
    return "".join(part[0].upper() for part in parts)[:3]


def cell_text(ws, row: int, col: int) -> str:
    value = ws.cell(row, col).value
    return "" if value is None else str(value).strip()


def main() -> None:
    wb = load_workbook(SOURCE, data_only=True)
    ws = wb["Sheet1"]

    stage_specs = [
        ("stage_july_15", "July 15 / Weeks 1-4", 0, "2026-07-15", range(5, 9)),
        ("stage_august_15", "August 15 / Weeks 5-8", 1, "2026-08-15", range(9, 13)),
        (
            "stage_september_15",
            "September 15 / Weeks 9-13",
            2,
            "2026-09-15",
            range(13, 18),
        ),
        (
            "stage_october_15",
            "October 15 / Weeks 14-17",
            3,
            "2026-10-15",
            range(18, 22),
        ),
        (
            "stage_november_15",
            "November 15 / Weeks 18-20",
            4,
            "2026-11-15",
            range(22, 25),
        ),
    ]
    col_to_stage = {col: spec for spec in stage_specs for col in spec[4]}
    stages = [
        {
            "id": stage_id,
            "name": name,
            "order": order,
            "colorToken": f"stage-{order + 1}",
            "rect": {"x": order * 320, "y": 0, "width": 300, "height": 2200},
        }
        for stage_id, name, order, _, _ in stage_specs
    ]

    category_rows: dict[int, str] = {}
    for row in range(1, ws.max_row + 1):
        category = cell_text(ws, row, 2)
        task = cell_text(ws, row, 3)
        if category and not task and category.upper() == category:
            category_rows[row] = category.title().replace(" + ", " + ")

    source_rows = []
    current_category = None
    current_owner = None
    for row in range(1, ws.max_row + 1):
        if row <= 3:
            continue

        if row in category_rows:
            current_category = category_rows[row]
            current_owner = None
            continue

        owner = cell_text(ws, row, 2)
        task = cell_text(ws, row, 3)
        if owner.lower() == "category" and task.lower() == "task":
            continue

        if owner and task:
            current_owner = owner
        if not task:
            continue

        markers = []
        budget_markers = []
        for col in range(5, 25):
            value = ws.cell(row, col).value
            if value is None:
                continue

            text_value = str(value).strip()
            if not text_value:
                continue

            label = cell_text(ws, 2, col) or f"Column {col}"
            if text_value.lower() == "x":
                markers.append((col, label))
            else:
                budget_markers.append((col, label, text_value))

        source_rows.append(
            {
                "row": row,
                "category": current_category or "General",
                "owner": current_owner or owner or "Unassigned",
                "task": task,
                "markers": markers,
                "budget_markers": budget_markers,
            }
        )

    infer_stage = {
        "Develop pitch deck": "stage_july_15",
        "Update business model + financials": "stage_july_15",
        "Comparisons to other Miami offerings": "stage_july_15",
        "Networking with Miami investors": "stage_july_15",
        "Draw up contract / terms w/ investor": "stage_august_15",
        "Receive cash in bank": "stage_august_15",
        "Define requirements / write spec": "stage_july_15",
        "Contact Miami realtors": "stage_july_15",
        "Visit sites / talk to landlords": "stage_july_15",
        "Zoning requirements thought through": "stage_july_15",
        "Open negotiations on 2+ sites": "stage_july_15",
        "Agree terms of lease": "stage_august_15",
        "Sign lease": "stage_august_15",
        "Decide 2-3 options for single piece of content": "stage_july_15",
        "Explore feasibility of each option": "stage_july_15",
        "Choose single piece to deliver": "stage_august_15",
        "IP rights licensing": "stage_august_15",
        "Detailed design plan": "stage_september_15",
        "Recruit creatives + technical team": "stage_september_15",
        "Set up studio": "stage_september_15",
        "Produce content": "stage_october_15",
        "On-site content tweaks": "stage_november_15",
        "Write design brief": "stage_july_15",
        "Contact + onboard architect": "stage_august_15",
        "Contact + onboard structural engineer": "stage_august_15",
        "Contact + onboard electrical engineer": "stage_august_15",
        "Design upgrade components": "stage_september_15",
        "Produce permit set": "stage_september_15",
        "Produce procurement documents": "stage_october_15",
        "Architectural design of ancillary spaces": "stage_september_15",
        "Produce permit set for ASs": "stage_october_15",
        "Produce procurement docs for ASs": "stage_october_15",
        "Retrieve containers to warehouse": "stage_august_15",
        "Review items + specify upgrades, etc.": "stage_august_15",
        "Contact suppliers w/ procurement docs": "stage_october_15",
        "Fabrication": "stage_october_15",
        "Shipping": "stage_november_15",
        "Hire installation team / rent equipment": "stage_november_15",
        "Installation": "stage_november_15",
        "Permit sign-off": "stage_november_15",
        "Commission AV system": "stage_november_15",
        "Fit-out of ancillary spaces": "stage_november_15",
        "Find ops partner": "stage_august_15",
        "Training": "stage_november_15",
        "A-B test ads for content type": "stage_august_15",
        "Produce marketing collateral": "stage_september_15",
        "Decide if we can bypass promoters": "stage_september_15",
        "Marketing": "stage_october_15",
        "Ticketing website": "stage_october_15",
        "Insurance": "stage_july_15",
        "Employee healthcare, etc.": "stage_september_15",
        "Bank account / other business finances": "stage_july_15",
        "Set up legal entity to isolate Miami from SSHQ": "stage_july_15",
        "Project management tool": "stage_july_15",
        "Google drive organization": "stage_july_15",
        "Taxes": "stage_august_15",
        "RVF": "stage_august_15",
        "Cap table": "stage_august_15",
        "Company structure": "stage_august_15",
        "Advisory agreements": "stage_august_15",
    }

    stage_date = {stage_id: date for stage_id, _, _, date, _ in stage_specs}
    stage_x = {stage["id"]: stage["rect"]["x"] for stage in stages}

    people_names = []
    for row in source_rows:
        if row["owner"] != "Unassigned" and row["owner"] not in people_names:
            people_names.append(row["owner"])
    people = [
        {
            "id": slug("person", name),
            "name": name,
            "initials": initials(name),
            "kind": "person",
        }
        for name in people_names
    ]

    categories = []
    for row in source_rows:
        if row["category"] not in categories:
            categories.append(row["category"])

    color_tokens = [
        "funding",
        "venue",
        "music",
        "structure",
        "sphere",
        "operations",
        "ticketing",
        "legal",
        "labs",
        "experience",
    ]
    workstreams = []
    for index, category in enumerate(categories):
        owner_ids = sorted(
            {
                slug("person", row["owner"])
                for row in source_rows
                if row["category"] == category and row["owner"] != "Unassigned"
            }
        )
        workstreams.append(
            {
                "id": slug("workstream", category),
                "name": category,
                "defaultAssociatedIds": owner_ids,
                "colorToken": color_tokens[index % len(color_tokens)],
            }
        )

    tags = [
        {
            "id": "tag_critical_path",
            "label": "Critical path",
            "colorToken": "critical",
        },
        {
            "id": "tag_funding_gate",
            "label": "Funding gate",
            "colorToken": "decision",
        },
        {
            "id": "tag_external_partner",
            "label": "External partner",
            "colorToken": "partner",
        },
        {
            "id": "tag_launch_readiness",
            "label": "Launch readiness",
            "colorToken": "experience",
        },
    ]

    critical_tasks = {
        "Receive cash in bank",
        "Sign lease",
        "Produce permit set",
        "Produce procurement documents",
        "Fabrication",
        "Shipping",
        "Installation",
        "Permit sign-off",
        "Commission AV system",
        "Produce content",
        "Ticketing website",
    }
    external_tasks = {
        task
        for task in infer_stage
        if any(
            word in task.lower()
            for word in [
                "contact",
                "onboard",
                "landlord",
                "realtor",
                "supplier",
                "partner",
                "investor",
            ]
        )
    }
    launch_tasks = {
        "Training",
        "Marketing",
        "Ticketing website",
        "Installation",
        "Permit sign-off",
        "Commission AV system",
        "On-site content tweaks",
    }
    funding_tasks = {
        row["task"] for row in source_rows if row["category"] == "Funding"
    }

    nodes = []
    y_by_stage = {stage["id"]: 90 for stage in stages}
    node_id_by_title = {}

    for row in source_rows:
        marker_stage = None
        if row["markers"]:
            first_col = min(col for col, _ in row["markers"])
            marker_stage = col_to_stage.get(first_col, (None,))[0]
        elif row["budget_markers"]:
            first_col = min(col for col, _, _ in row["budget_markers"])
            marker_stage = col_to_stage.get(first_col, (None,))[0]

        stage_id = marker_stage or infer_stage.get(row["task"], "stage_august_15")
        notes = [f"Source row {row['row']}: {row['category']}"]
        if row["markers"]:
            notes.append(
                "Timeline marker(s): "
                + ", ".join(label for _, label in row["markers"])
            )
        if row["budget_markers"]:
            notes.append(
                "Budget marker(s): "
                + ", ".join(
                    f"{value} at {label}"
                    for _, label, value in row["budget_markers"]
                )
            )
        if not marker_stage:
            notes.append(
                "Stage inferred from task order and delivery dependency sequence."
            )

        tag_ids = []
        if row["task"] in critical_tasks:
            tag_ids.append("tag_critical_path")
        if row["task"] in funding_tasks:
            tag_ids.append("tag_funding_gate")
        if row["task"] in external_tasks:
            tag_ids.append("tag_external_partner")
        if row["task"] in launch_tasks:
            tag_ids.append("tag_launch_readiness")

        priority = (
            "P0"
            if row["task"] in critical_tasks
            else "P1"
            if row["category"] in {"Funding", "Site", "Architecture", "Delivery"}
            else "P2"
        )
        node_id = slug("node", row["task"])
        base_id = node_id
        suffix = 2
        while node_id in node_id_by_title.values():
            node_id = f"{base_id}_{suffix}"
            suffix += 1
        node_id_by_title[row["task"]] = node_id

        y = y_by_stage[stage_id]
        y_by_stage[stage_id] += 155
        nodes.append(
            {
                "id": node_id,
                "type": "planningNode",
                "position": {"x": stage_x[stage_id] + 30, "y": y},
                "data": {
                    "title": row["task"],
                    "status": "defined"
                    if marker_stage or row["budget_markers"]
                    else "idea",
                    "associatedIds": []
                    if row["owner"] == "Unassigned"
                    else [slug("person", row["owner"])],
                    "workstreamId": slug("workstream", row["category"]),
                    "tagIds": tag_ids,
                    "stageId": stage_id,
                    "priority": priority,
                    "targetDate": stage_date[stage_id],
                    "confidence": "medium"
                    if marker_stage or row["budget_markers"]
                    else "low",
                    "notes": "\n".join(notes),
                    "createdAt": "2026-06-12T00:00:00.000Z",
                    "updatedAt": "2026-06-12T00:00:00.000Z",
                },
            }
        )

    band_nodes = [
        {
            "id": f"tab_miami_delivery_strategy_band_{stage['id']}",
            "type": "stageBand",
            "position": {"x": stage["rect"]["x"], "y": stage["rect"]["y"]},
            "width": stage["rect"]["width"],
            "height": stage["rect"]["height"],
            "data": {
                "title": stage["name"],
                "stageId": stage["id"],
                "locked": True,
                "colorToken": stage["colorToken"],
            },
            "draggable": False,
            "selectable": False,
            "deletable": False,
            "zIndex": 0,
        }
        for stage in stages
    ]

    edges = []
    edge_seen = set()

    def add_edge(source_title: str, target_title: str, line: str = "solid") -> None:
        source_id = node_id_by_title.get(source_title)
        target_id = node_id_by_title.get(target_title)
        if not source_id or not target_id or source_id == target_id:
            return
        key = (source_id, target_id)
        if key in edge_seen:
            return
        edge_seen.add(key)
        edges.append(
            {
                "id": f"edge_{source_id}_to_{target_id}",
                "source": source_id,
                "target": target_id,
                "type": "planningEdge",
                "data": {"lineType": line},
            }
        )

    for category in categories:
        category_rows_in_order = [
            row for row in source_rows if row["category"] == category
        ]
        for source_row, target_row in zip(
            category_rows_in_order, category_rows_in_order[1:]
        ):
            add_edge(source_row["task"], target_row["task"])

    cross_edges = [
        ("Receive cash in bank", "Sign lease"),
        ("Sign lease", "Produce permit set"),
        ("Produce permit set", "Produce procurement documents"),
        ("Produce procurement documents", "Contact suppliers w/ procurement docs"),
        ("Contact suppliers w/ procurement docs", "Fabrication"),
        ("Fabrication", "Shipping"),
        ("Shipping", "Installation"),
        ("Installation", "Permit sign-off"),
        ("Permit sign-off", "Commission AV system"),
        ("Choose single piece to deliver", "A-B test ads for content type"),
        ("Choose single piece to deliver", "IP rights licensing"),
        ("Detailed design plan", "Produce content"),
        ("Produce content", "On-site content tweaks"),
        ("Produce marketing collateral", "Marketing"),
        ("Decide if we can bypass promoters", "Ticketing website"),
        ("Find ops partner", "Training"),
        ("Insurance", "Sign lease"),
        ("Set up legal entity to isolate Miami from SSHQ", "Sign lease"),
        ("Project management tool", "Google drive organization"),
    ]
    for source, target in cross_edges:
        add_edge(source, target, "magic")

    project = {
        "schemaVersion": 1,
        "projectName": "SS Miami Delivery Strategy",
        "activeTabId": "tab_miami_delivery_strategy",
        "people": people,
        "workstreams": workstreams,
        "tags": tags,
        "tabs": [
            {
                "id": "tab_miami_delivery_strategy",
                "name": "Miami Delivery Strategy",
                "description": (
                    "Imported from SS Miami delivery strategy.xlsx. Stage placement "
                    "uses visible week markers where present and inferred delivery "
                    "sequence where the source row had no marker."
                ),
                "orientation": "vertical",
                "stages": stages,
                "nodes": band_nodes + nodes,
                "edges": edges,
                "viewport": {"x": -40, "y": -20, "zoom": 0.65},
                "filters": {},
            }
        ],
        "snapshots": [],
        "settings": {
            "themeId": "clean-light",
            "showMiniMap": True,
            "adminMode": False,
            "presentationMode": False,
        },
        "createdAt": "2026-06-12T00:00:00.000Z",
        "updatedAt": "2026-06-12T00:00:00.000Z",
    }

    planning_ids = {node["id"] for node in nodes}
    for edge in edges:
        assert edge["source"] in planning_ids, edge
        assert edge["target"] in planning_ids, edge

    OUTPUT.write_text(json.dumps(project, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(OUTPUT),
                "tasks": len(nodes),
                "edges": len(edges),
                "people": len(people),
                "workstreams": len(workstreams),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
